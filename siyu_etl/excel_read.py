"""
Excel 数据读取模块

该模块负责：
1. 从 Excel 文件中流式读取数据行
2. 处理表头行识别
3. 在遇到"合计"/"总计"行或连续空行时停止读取
4. 返回清洗后的数据行
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterator

import openpyxl

from siyu_etl.cleaner import clean_row
from siyu_etl.constants import MAX_SCAN_COLS, MAX_SCAN_ROWS


@dataclass(frozen=True)
class ReadRow:
    """
    读取的行数据
    
    Attributes:
        row_index_0based: 行索引（0-based）
        data: 清洗后的数据字典（所有值都是字符串）
        raw: 原始数据字典（包含原始类型）
    """
    row_index_0based: int
    data: dict[str, str]
    raw: dict[str, Any]


def read_rows(
    file_path: Path,
    *,
    header_row_0based: int,
    headers: list[str] | None = None,
    stop_on_total_keywords: tuple[str, ...] = ("合计", "总计"),
    max_consecutive_empty_rows: int = 3,
    file_type: str | None = None,
) -> Iterator[ReadRow]:
    """
    流式读取 Excel 文件的数据行
    
    功能：
    - 所有输出值都是字符串
    - 遇到"合计/总计"行或连续空行时停止读取
    
    Args:
        file_path: Excel 文件路径
        header_row_0based: 表头行索引（0-based）
        headers: 预填充的表头列表（来自 fill_header_row）。如果为 None，将从工作表中读取
        stop_on_total_keywords: 停止读取的关键词元组（默认："合计"、"总计"）
        max_consecutive_empty_rows: 最大连续空行数，超过此数量则停止读取（默认：3）
        file_type: 文件类型（用于数据清洗时的特殊处理，如百分比字段转换）
        
    Yields:
        ReadRow 对象，包含清洗后的行数据
    """
    file_path = Path(file_path)
    # NOTE: Some exported workbooks have broken dimensions (max_row/max_column=1).
    # Force a reasonable scan window; we will stop early on empty streak / total row.
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    except PermissionError as e:
        raise PermissionError(
            f"无法读取文件 {file_path.name}：文件可能正在被其他程序（如 Excel）打开。"
            f"请关闭文件后重试。原始错误: {e}"
        ) from e
    except Exception as e:
        raise RuntimeError(f"读取 Excel 文件失败: {file_path.name}。错误: {e}") from e
    try:
        ws = wb.worksheets[0]

        header_row_idx = header_row_0based + 1  # openpyxl is 1-based
        scan_max_col = MAX_SCAN_COLS
        
        # Use provided headers if available, otherwise read from sheet
        if headers is None:
            header_cells = next(
                ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx, max_col=scan_max_col)
            )
            headers = []
            for c in header_cells:
                h = str(c.value).strip() if c.value is not None else ""
                headers.append(h)

            # Remove trailing empty headers but keep alignment for column access
            last_non_empty = 0
            for i, h in enumerate(headers):
                if h:
                    last_non_empty = i
            headers = headers[: last_non_empty + 1]

        empty_streak = 0
        scan_max_row = header_row_idx + MAX_SCAN_ROWS
        for row_idx_1based, row_cells in enumerate(
            ws.iter_rows(
                min_row=header_row_idx + 1,
                max_row=scan_max_row,
                max_col=max(len(headers), 1),
            ),
            start=header_row_idx + 1,
        ):
            row_cells = list(row_cells)[: len(headers)]

            raw: dict[str, Any] = {}
            number_formats: dict[str, str] = {}
            row_text_concat = ""
            all_empty = True

            for h, cell in zip(headers, row_cells, strict=False):
                if not h:
                    continue
                v = getattr(cell, "value", None)
                raw[h] = v
                number_formats[h] = getattr(cell, "number_format", "") or ""
                if v is not None and str(v).strip() != "":
                    all_empty = False
                row_text_concat += str(v or "")

            # Stop conditions
            if any(k in row_text_concat for k in stop_on_total_keywords):
                break

            if all_empty:
                empty_streak += 1
                if empty_streak >= max_consecutive_empty_rows:
                    break
                continue
            empty_streak = 0

            cleaned = clean_row(raw, number_formats=number_formats, file_type=file_type)
            yield ReadRow(
                row_index_0based=row_idx_1based - 1,
                data=cleaned.data,
                raw=raw,
            )
    finally:
        wb.close()


