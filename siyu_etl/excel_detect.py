"""
Excel 文件类型检测和表头识别模块

该模块负责：
1. 根据文件名识别 Excel 文件的类型（会员交易明细、店内订单明细等）
2. 检测并填充表头行（处理合并单元格、多行表头等情况）
3. 返回检测结果，包括文件类型、表头行位置、表头列表等信息
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable, Optional

import openpyxl

from siyu_etl.cleaner import normalize_cell_to_string
from siyu_etl.constants import MAX_SCAN_COLS


# 文件类型常量定义
FILETYPE_MEMBER_TRADE = "会员交易明细"
FILETYPE_INSTORE_ORDER = "店内订单明细(已结账)"
FILETYPE_INCOME_DISCOUNT = "收入优惠统计"
FILETYPE_COUPON_STAT = "优惠券统计表"
FILETYPE_MEMBER_STORAGE = "会员储值消费分析表"
FILETYPE_MEMBER_CARD_EXPORT = "会员卡导出"


@dataclass(frozen=True)
class TableSpec:
    """
    表规格定义，用于描述每种文件类型的特征
    
    Attributes:
        file_type: 文件类型名称
        filename_keywords: 文件名关键词元组，用于识别文件类型
        required_headers: 必需的表头字段元组
        default_header_row_0based: 默认表头行索引（0-based）
        timestamp_column: 时间戳列名
    """
    file_type: str
    filename_keywords: tuple[str, ...]
    required_headers: tuple[str, ...]
    default_header_row_0based: int
    timestamp_column: str


SPECS: tuple[TableSpec, ...] = (
    TableSpec(
        file_type=FILETYPE_MEMBER_TRADE,
        filename_keywords=("会员交易明细",),
        required_headers=("交易流水号", "交易时间"),
        default_header_row_0based=1,
        timestamp_column="交易时间",
    ),
    TableSpec(
        file_type=FILETYPE_INSTORE_ORDER,
        filename_keywords=("店内订单明细", "订单明细"),
        required_headers=("下单时间",),
        default_header_row_0based=2,
        timestamp_column="下单时间",
    ),
    TableSpec(
        file_type=FILETYPE_INCOME_DISCOUNT,
        filename_keywords=("收入优惠统计",),
        required_headers=("门店", "营业日期", "结账方式"),
        default_header_row_0based=2,
        timestamp_column="营业日期",
    ),
    TableSpec(
        file_type=FILETYPE_COUPON_STAT,
        filename_keywords=("优惠券统计表", "优惠券统计"),
        required_headers=("交易日期", "券名称", "券类型"),
        default_header_row_0based=2,
        timestamp_column="交易日期",
    ),
    TableSpec(
        file_type=FILETYPE_MEMBER_STORAGE,
        filename_keywords=("会员储值消费分析表", "储值消费分析"),
        required_headers=("交易日期", "机构编码", "开卡门店"),
        default_header_row_0based=3,
        timestamp_column="交易日期",
    ),
    TableSpec(
        file_type=FILETYPE_MEMBER_CARD_EXPORT,
        filename_keywords=("会员卡导出",),
        required_headers=("会员卡号", "开卡时间", "开卡门店"),
        default_header_row_0based=1,
        timestamp_column="开卡时间",
    ),
)


@dataclass(frozen=True)
class DetectedSheet:
    """
    检测结果数据类
    
    Attributes:
        file_path: Excel 文件路径
        file_type: 识别出的文件类型
        header_row_0based: 表头行索引（0-based）
        headers: 表头列表
        timestamp_column: 时间戳列名
    """
    file_path: Path
    file_type: str
    header_row_0based: int
    headers: list[str]
    timestamp_column: str


def _guess_by_filename(p: Path) -> Optional[TableSpec]:
    """
    根据文件名猜测文件类型
    
    Args:
        p: 文件路径
        
    Returns:
        匹配的 TableSpec，如果无法识别则返回 None
    """
    name = p.name
    for spec in SPECS:
        if any(k in name for k in spec.filename_keywords):
            return spec
    return None


def _row_values_as_str(row_cells: Iterable) -> list[str]:
    """
    将行单元格转换为字符串列表
    
    Args:
        row_cells: 单元格迭代器
        
    Returns:
        字符串列表，每个元素对应一个单元格的值（已标准化）
    """
    vals: list[str] = []
    for c in row_cells:
        try:
            fmt = getattr(c, "number_format", None)
            vals.append(normalize_cell_to_string(getattr(c, "value", None), number_format=fmt))
        except Exception:
            vals.append("")
    return [v.strip() for v in vals]


def _score_header_row(values: list[str], required: tuple[str, ...]) -> int:
    """
    计算表头行的匹配分数（匹配的必需字段数量）
    
    Args:
        values: 当前行的值列表
        required: 必需的字段元组
        
    Returns:
        匹配的必需字段数量
    """
    s = set(v for v in values if v)
    return sum(1 for r in required if r in s)


def detect_sheet(file_path: Path, scan_rows: int = 20) -> DetectedSheet:
    """
    检测 Excel 文件的表类型和表头行
    
    根据 PRD 要求：表头行位置直接指定，无需自动检测。
    该函数会：
    1. 根据文件名识别文件类型
    2. 获取指定的表头行位置
    3. 填充表头（处理合并单元格和多行表头）
    
    Args:
        file_path: Excel 文件路径
        scan_rows: 扫描行数（当前未使用，保留用于兼容性）
        
    Returns:
        DetectedSheet 对象，包含检测结果
        
    Raises:
        ValueError: 无法识别文件类型时抛出
    """
    file_path = Path(file_path)
    try:
        # 注意：read_only模式下无法读取合并单元格信息（merged_cells属性不存在）
        # 但我们需要合并单元格信息来正确处理表头，所以不能使用read_only模式
        # 使用keep_vba=False来避免加载VBA代码，提高性能
        wb = openpyxl.load_workbook(file_path, read_only=False, data_only=True, keep_vba=False)
    except PermissionError as e:
        raise PermissionError(
            f"无法读取文件 {file_path.name}：文件可能正在被其他程序（如 Excel）打开。"
            f"请关闭文件后重试。原始错误: {e}"
        ) from e
    except Exception as e:
        raise RuntimeError(f"读取 Excel 文件失败: {file_path.name}。错误: {e}") from e
    
    try:
        ws = wb.worksheets[0]
        scan_max_col = MAX_SCAN_COLS

        # 1) Identify file type by filename
        spec = _guess_by_filename(file_path)
        if spec is None:
            supported_types = "、".join([
                "会员交易明细", "店内订单明细(已结账)", "收入优惠统计",
                "优惠券统计表", "会员储值消费分析表", "会员卡导出"
            ])
            raise ValueError(
                f"无法识别表类型: {file_path.name}\n"
                f"支持的文件类型：{supported_types}。\n"
                f"请确认文件名包含上述关键词之一。"
            )

        # 2) Get specified header row (1-based)
        header_row_1based = get_header_row_for_file_type(spec.file_type)
        header_row_0based = header_row_1based - 1

        # 3) Fill header row
        # 规则：
        # - 数据区是表头行的下一行开始
        # - 对于数据区上方的区域（包括表头行），如果单元格在合并区域内，取合并区域左上角的值
        # - 以指定的表头行作为表头，不进行向上查找和向下合并
        next_row_1based = header_row_1based - 1 if spec.file_type == FILETYPE_MEMBER_STORAGE else None
        headers = fill_header_row(ws, header_row_1based, next_row_1based=next_row_1based, max_col=scan_max_col)
        
        # Remove empty headers but keep alignment
        headers = [h for h in headers if h]

        return DetectedSheet(
            file_path=file_path,
            file_type=spec.file_type,
            header_row_0based=header_row_0based,
            headers=headers,
            timestamp_column=spec.timestamp_column,
        )
    except ValueError:
        # 重新抛出 ValueError（文件类型识别失败）
        raise
    except Exception as e:
        # 其他异常包装为更友好的错误信息
        raise RuntimeError(
            f"处理 Excel 文件失败: {file_path.name}。"
            f"可能原因：文件格式不正确、表头识别失败等。错误: {e}"
        ) from e
    finally:
        wb.close()

def get_header_row_for_file_type(file_type: str) -> int:
    """
    获取指定文件类型的表头行位置（1-based）
    
    根据 PRD 要求：表头行位置直接指定，无需自动检测。
    
    Args:
        file_type: 文件类型
        
    Returns:
        表头行位置（1-based），如果未找到则返回默认值 2
    """
    mapping = {
        FILETYPE_MEMBER_TRADE: 2,  # Row 2 (1-based)
        FILETYPE_INSTORE_ORDER: 3,  # Row 3 (1-based)
        FILETYPE_INCOME_DISCOUNT: 3,  # Row 3 (1-based)
        FILETYPE_COUPON_STAT: 3,  # Row 3 (1-based)
        FILETYPE_MEMBER_STORAGE: 4,  # Row 4 (1-based)
        FILETYPE_MEMBER_CARD_EXPORT: 2,  # Row 2 (1-based)
    }
    return mapping.get(file_type, 2)


def split_merged_cells(ws, row_idx: int, col_idx: int) -> Optional[Any]:
    """
    拆分合并单元格：如果单元格是合并区域的一部分，返回合并区域左上角值；
    若工作表无法提供 merged_cells 信息，则对同一行做一次轻量向左回溯，尽量还原常见横向合并表头。
    """
    cell_value = ws.cell(row=row_idx, column=col_idx).value
    if hasattr(ws, "merged_cells"):
        merged_ranges = list(ws.merged_cells.ranges)
        for merged_range in merged_ranges:
            if (
                merged_range.min_row <= row_idx <= merged_range.max_row
                and merged_range.min_col <= col_idx <= merged_range.max_col
            ):
                top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                return top_left_cell.value
        return cell_value

    if cell_value is not None:
        return cell_value

    for left_col in range(col_idx - 1, 0, -1):
        left_val = ws.cell(row=row_idx, column=left_col).value
        if left_val is not None:
            return left_val
    return None


def fill_header_row(
    ws, header_row_1based: int, next_row_1based: int | None = None, max_col: int = MAX_SCAN_COLS
) -> list[str]:
    """
    填充表头行，处理合并单元格和双层表头。

    规则：
    1. 默认读取指定表头行。
    2. 若提供 next_row_1based，则把下一行视为补充表头：
       - 主表头为空且子表头有值时，直接采用子表头
       - 主表头与子表头同时有值时，拼接为一个完整表头
    3. 保留中间空列位置，仅去掉尾部连续空列。
    
    Args:
        ws: openpyxl 工作表对象
        header_row_1based: 表头行位置（1-based）
        max_col: 最大扫描列数
        
    Returns:
        表头字符串列表
    """
    header_row = list(ws.iter_rows(min_row=header_row_1based, max_row=header_row_1based, max_col=max_col))[0]
    header_values: list[str] = []

    for col_idx, cell in enumerate(header_row, start=1):
        val = cell.value
        header_values.append(str(val).strip() if val is not None else "")

    if next_row_1based is not None:
        next_row = list(ws.iter_rows(min_row=next_row_1based, max_row=next_row_1based, max_col=max_col))[0]
        max_cols = max(len(header_values), len(next_row))
        while len(header_values) < max_cols:
            header_values.append("")

        for col_idx in range(1, max_cols + 1):
            next_cell = next_row[col_idx - 1] if col_idx - 1 < len(next_row) else None
            next_val = str(next_cell.value).strip() if next_cell is not None and next_cell.value is not None else ""

            current_val = header_values[col_idx - 1]
            if not current_val and next_val:
                header_values[col_idx - 1] = next_val
            elif current_val and next_val and next_val not in current_val:
                header_values[col_idx - 1] = f"{current_val}{next_val}"

    while header_values and not header_values[-1]:
        header_values.pop()

    return header_values



def _infer_header_row_0based(ws, header_values: list[str], scan_rows: int, max_col: int) -> int:
    """
    推断表头行位置（0-based）
    
    当多行有相似匹配时，通过严格相等匹配推断确切的行。
    如果找不到完全匹配，则回退到扫描行中第一个匹配的行。
    
    Args:
        ws: openpyxl 工作表对象
        header_values: 目标表头值列表
        scan_rows: 扫描行数
        max_col: 最大扫描列数
        
    Returns:
        表头行索引（0-based），如果找不到则返回 0
    """
    target = [v.strip() for v in header_values]
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=scan_rows, max_col=max_col), start=1
    ):
        values = _row_values_as_str(row)
        if values == target:
            return row_idx - 1
    # fallback: assume header is within scan rows, take first non-empty row with the same set.
    target_set = set(v for v in target if v)
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=scan_rows, max_col=max_col), start=1
    ):
        values = _row_values_as_str(row)
        if target_set.issubset(set(values)):
            return row_idx - 1
    return 0


