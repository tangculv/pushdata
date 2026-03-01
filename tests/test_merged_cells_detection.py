#!/usr/bin/env python3
"""
检查A-G列的跨行合并单元格检测问题
"""

import openpyxl
from pathlib import Path

def main():
    project_root = Path(__file__).resolve().parents[1]
    excel_file = project_root / "data" / "山禾田_会员储值消费分析表_2026-01-24 12_17_45_a049681_1769228269877.xlsx"
    
    print("=" * 100)
    print("检查A-G列的跨行合并单元格")
    print("=" * 100)
    
    # 尝试两种模式
    for read_only in [True, False]:
        print(f"\n{'='*100}")
        print(f"模式: read_only={read_only}")
        print("-" * 100)
        
        try:
            wb = openpyxl.load_workbook(excel_file, read_only=read_only, data_only=True)
            ws = wb.worksheets[0]
            
            # 检查是否有merged_cells属性
            if not hasattr(ws, "merged_cells"):
                print("工作表没有merged_cells属性")
            else:
                merged_ranges = list(ws.merged_cells.ranges)
                print(f"找到 {len(merged_ranges)} 个合并单元格区域")
                
                if merged_ranges:
                    print("\n所有合并单元格区域:")
                    for merged_range in merged_ranges:
                        min_col_letter = openpyxl.utils.get_column_letter(merged_range.min_col)
                        max_col_letter = openpyxl.utils.get_column_letter(merged_range.max_col)
                        min_row_val = ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
                        val_str = str(min_row_val).strip() if min_row_val is not None else "(空)"
                        print(f"  行{merged_range.min_row}-{merged_range.max_row}, 列{merged_range.min_col}-{merged_range.max_col} ({min_col_letter}-{max_col_letter}): '{val_str}'")
                else:
                    print("没有找到合并单元格")
            
            # 检查A-G列（列1-7）是否有跨行合并（涉及第3行和第4行）
            print("\n检查A-G列（列1-7）的跨行合并:")
            print("-" * 100)
            for col_idx in range(1, 8):
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                val3 = ws.cell(row=3, column=col_idx).value
                val4 = ws.cell(row=4, column=col_idx).value
                val3_str = str(val3).strip() if val3 is not None else "(空)"
                val4_str = str(val4).strip() if val4 is not None else "(空)"
                
                # 检查是否在合并区域内
                in_merged = False
                merged_val = None
                if hasattr(ws, "merged_cells"):
                    for merged_range in ws.merged_cells.ranges:
                        if (
                            merged_range.min_row <= 4 <= merged_range.max_row
                            and merged_range.min_col <= col_idx <= merged_range.max_col
                            and merged_range.min_row < merged_range.max_row  # 跨行
                        ):
                            in_merged = True
                            merged_val = ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
                            min_col_letter = openpyxl.utils.get_column_letter(merged_range.min_col)
                            max_col_letter = openpyxl.utils.get_column_letter(merged_range.max_col)
                            print(f"  列{col_letter} (列{col_idx}): 在合并区域内 (行{merged_range.min_row}-{merged_range.max_row}, 列{min_col_letter}-{max_col_letter})")
                            print(f"    第3行值: '{val3_str}'")
                            print(f"    第4行值: '{val4_str}'")
                            print(f"    合并区域左上角值: '{merged_val}'")
                            print(f"    -> 第4行应该填充为: '{merged_val}'")
                            break
                
                if not in_merged:
                    print(f"  列{col_letter} (列{col_idx}): 不在合并区域内")
                    print(f"    第3行值: '{val3_str}'")
                    print(f"    第4行值: '{val4_str}'")
            
            wb.close()
            
        except Exception as e:
            print(f"错误: {e}")
            import traceback
            traceback.print_exc()

if __name__ == "__main__":
    main()
