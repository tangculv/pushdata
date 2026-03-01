from __future__ import annotations

from pathlib import Path

import openpyxl

from siyu_etl.excel_detect import (
    FILETYPE_MEMBER_STORAGE,
    detect_sheet,
    fill_header_row,
    get_header_row_for_file_type,
    split_merged_cells,
)


def test_get_header_row_for_file_type() -> None:
    """Test header row mapping."""
    from siyu_etl.excel_detect import (
        FILETYPE_COUPON_STAT,
        FILETYPE_INCOME_DISCOUNT,
        FILETYPE_INSTORE_ORDER,
        FILETYPE_MEMBER_STORAGE,
        FILETYPE_MEMBER_TRADE,
    )

    assert get_header_row_for_file_type(FILETYPE_MEMBER_TRADE) == 2
    assert get_header_row_for_file_type(FILETYPE_INSTORE_ORDER) == 3
    assert get_header_row_for_file_type(FILETYPE_INCOME_DISCOUNT) == 3
    assert get_header_row_for_file_type(FILETYPE_COUPON_STAT) == 3
    assert get_header_row_for_file_type(FILETYPE_MEMBER_STORAGE) == 4


def test_split_merged_cells(tmp_path: Path) -> None:
    """Test merged cell splitting."""
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Create merged cells: A1:C1
    ws["A1"] = "合并表头"
    ws.merge_cells("A1:C1")
    ws["D1"] = "其他列"
    
    path = tmp_path / "test_merged.xlsx"
    wb.save(path)
    wb.close()
    
    # Reload to test
    wb2 = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws2 = wb2.worksheets[0]
    
    # Test splitting merged cells
    val_a = split_merged_cells(ws2, 1, 1)  # A1 (top-left)
    val_b = split_merged_cells(ws2, 1, 2)  # B1 (merged, should get A1 value)
    val_c = split_merged_cells(ws2, 1, 3)  # C1 (merged, should get A1 value)
    val_d = split_merged_cells(ws2, 1, 4)  # D1 (not merged)
    
    assert val_a == "合并表头"
    assert val_b == "合并表头"  # Should get top-left value
    assert val_c == "合并表头"  # Should get top-left value
    assert val_d == "其他列"
    
    wb2.close()


def test_fill_header_row_multirow(tmp_path: Path) -> None:
    """Test multi-row header filling (like 会员储值消费分析表)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Row 3: main header with some None values
    ws.append(["说明行"])
    ws.append(["说明行2"])
    row3 = ["交易日期", "卡类型名称", "机构编码", "累计储值金额（元）", None, None, "会员余额（元）"]
    ws.append(row3)
    # Row 4: supplementary header
    row4 = [None, None, None, "储值余额累计（元）", "赠送余额累计（元）", "合计（元）", "储值余额（元）"]
    ws.append(row4)
    
    path = tmp_path / "test_multirow.xlsx"
    wb.save(path)
    wb.close()
    
    # Reload to test
    wb2 = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws2 = wb2.worksheets[0]
    
    headers = fill_header_row(ws2, header_row_1based=3, next_row_1based=4, max_col=10)
    
    # Check merged headers
    assert "交易日期" in headers
    assert "卡类型名称" in headers
    assert "机构编码" in headers
    # Col 4: should be merged
    assert any("累计储值金额（元）" in h and "储值余额累计（元）" in h for h in headers)
    # Col 5: should be from row 4
    assert "赠送余额累计（元）" in headers
    # Col 7: should be merged
    assert any("会员余额（元）" in h and "储值余额（元）" in h for h in headers)
    
    wb2.close()


def test_detect_member_storage_sheet(tmp_path: Path) -> None:
    """Test detection of 会员储值消费分析表 with multi-row header."""
    wb = openpyxl.Workbook()
    ws = wb.active
    
    ws.append(["会员储值消费分析表"])
    ws.append(["日期范围：【按日】；2025/12/01 至 2025/12/31"])
    # Row 3: main header
    ws.append(["交易日期", "卡类型名称", "机构编码", "开卡门店", "会员数量", "累计储值金额（元）", None, None, "会员余额（元）"])
    # Row 4: supplementary header
    ws.append([None, None, None, None, None, "储值余额累计（元）", "赠送余额累计（元）", "合计（元）", "储值余额（元）"])
    # Row 5: data
    ws.append(["2025/12/31", "山禾田会员卡", "MD00003", "山禾田·日料小屋（福田店）", 43720.0, 1865812.73, 152280.27, 2018093.0, 182271.13])
    
    path = tmp_path / "山禾田_会员储值消费分析表_测试.xlsx"
    wb.save(path)
    wb.close()
    
    det = detect_sheet(path)
    assert det.file_type == FILETYPE_MEMBER_STORAGE
    assert det.header_row_0based == 3  # Row 4 (1-based) = Row 3 (0-based)
    
    # Check that headers are filled and merged
    assert "交易日期" in det.headers
    assert "机构编码" in det.headers
    # Check merged headers
    assert any("累计储值金额（元）" in h for h in det.headers)
    assert any("储值余额累计（元）" in h for h in det.headers)

