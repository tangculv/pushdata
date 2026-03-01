#!/usr/bin/env python3
"""
完整分析表头结构：第3行和第4行的内容
"""

import openpyxl
from pathlib import Path

def get_merged_value(ws, row, col):
    """获取合并单元格的值"""
    if not hasattr(ws, "merged_cells"):
        return None
    for merged_range in ws.merged_cells.ranges:
        if (
            merged_range.min_row <= row <= merged_range.max_row
            and merged_range.min_col <= col <= merged_range.max_col
        ):
            return ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
    return None

def main():
    project_root = Path(__file__).resolve().parents[1]
    excel_file = project_root / "data" / "山禾田_会员储值消费分析表_2026-01-24 12_17_45_a049681_1769228269877.xlsx"
    
    wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    
    print("=" * 100)
    print("完整表头结构分析：逐列查看第3行和第4行的内容")
    print("=" * 100)
    
    row3 = list(ws.iter_rows(min_row=3, max_row=3, max_col=30))[0]
    row4 = list(ws.iter_rows(min_row=4, max_row=4, max_col=30))[0]
    
    print("\n逐列分析（列1-21）:")
    print("-" * 100)
    
    for col_idx in range(1, 22):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        
        # 第3行的值
        val3_merged = get_merged_value(ws, 3, col_idx)
        val3 = ws.cell(row=3, column=col_idx).value
        val3_final = val3_merged if val3_merged is not None else val3
        val3_str = str(val3_final).strip() if val3_final else ""
        
        # 第4行的值
        val4_merged = get_merged_value(ws, 4, col_idx)
        val4 = ws.cell(row=4, column=col_idx).value
        val4_final = val4_merged if val4_merged is not None else val4
        val4_str = str(val4_final).strip() if val4_final else ""
        
        # 确定表头应该是什么
        if val3_str and val4_str:
            # 两行都有值：用户说第4行是表头行，所以应该只用第4行的值
            # 但第3行的值可能是分类标题，不应该合并
            header = val4_str
            note = f"（第3行有'{val3_str}'，但表头应该是第4行的'{val4_str}'）"
        elif val3_str and not val4_str:
            # 只有第3行有值：使用第3行的值
            header = val3_str
            note = "（只有第3行有值）"
        elif not val3_str and val4_str:
            # 只有第4行有值：使用第4行的值
            header = val4_str
            note = "（只有第4行有值）"
        else:
            header = ""
            note = "（空）"
        
        if header:
            print(f"列{col_idx:2d} ({col_letter:2s}): '{header}' {note}")
    
    print("\n" + "=" * 100)
    print("总结：")
    print("-" * 100)
    print("根据用户要求：")
    print("1. 表头行是第4行")
    print("2. 第3行的'累计储值金额（元）'不应该合并到表头中")
    print("3. 对于第4行为空的列（如A-G列），应该使用第3行的值作为表头")
    print("4. 对于第4行有值的列（如H列及以后），应该只使用第4行的值，不合并第3行")
    
    wb.close()

if __name__ == "__main__":
    main()
