#!/usr/bin/env python3
"""
调试脚本：详细查看会员储值消费分析表的表头结构
"""

import openpyxl
from pathlib import Path

def main():
    project_root = Path(__file__).resolve().parents[1]
    excel_file = project_root / "data" / "山禾田_会员储值消费分析表_2026-01-24 12_17_45_a049681_1769228269877.xlsx"
    
    if not excel_file.exists():
        print(f"错误：文件不存在: {excel_file}")
        return
    
    wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    
    print("=" * 100)
    print("详细表头分析")
    print("=" * 100)
    
    # 检查合并单元格
    print("\n合并单元格区域:")
    if hasattr(ws, "merged_cells"):
        for merged_range in ws.merged_cells.ranges:
            if merged_range.min_row <= 4 and merged_range.max_row >= 3:
                print(f"  行{merged_range.min_row}-{merged_range.max_row}, 列{merged_range.min_col}-{merged_range.max_col}")
    
    # 读取第3行（主表头）
    print("\n第3行（主表头）- 所有列:")
    print("-" * 100)
    row3 = list(ws.iter_rows(min_row=3, max_row=3, max_col=30))[0]
    for col_idx, cell in enumerate(row3, start=1):
        val = cell.value
        if val is not None:
            print(f"  列{col_idx} ({openpyxl.utils.get_column_letter(col_idx)}): '{val}'")
    
    # 读取第4行（子表头）
    print("\n第4行（子表头）- 所有列:")
    print("-" * 100)
    row4 = list(ws.iter_rows(min_row=4, max_row=4, max_col=30))[0]
    for col_idx, cell in enumerate(row4, start=1):
        val = cell.value
        if val is not None:
            print(f"  列{col_idx} ({openpyxl.utils.get_column_letter(col_idx)}): '{val}'")
    
    # 逐列分析，显示合并后的表头
    print("\n逐列分析（合并后的表头）:")
    print("-" * 100)
    max_col = max(len(row3), len(row4))
    
    def get_merged_value(ws, row, col):
        """获取合并单元格的值"""
        if not hasattr(ws, "merged_cells"):
            return None
        for merged_range in ws.merged_cells.ranges:
            if (
                merged_range.min_row <= row <= merged_range.max_row
                and merged_range.min_col <= col <= merged_range.max_col
            ):
                return ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
        return None
    
    for col_idx in range(1, max_col + 1):
        # 获取第3行的值（检查合并单元格）
        val3_merged = get_merged_value(ws, 3, col_idx)
        val3 = ws.cell(row=3, column=col_idx).value
        val3_final = val3_merged if val3_merged is not None else val3
        
        # 获取第4行的值（检查合并单元格）
        val4_merged = get_merged_value(ws, 4, col_idx)
        val4 = ws.cell(row=4, column=col_idx).value
        val4_final = val4_merged if val4_merged is not None else val4
        
        # 合并显示
        if val3_final or val4_final:
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            val3_str = str(val3_final).strip() if val3_final else ""
            val4_str = str(val4_final).strip() if val4_final else ""
            
            if val3_str and val4_str:
                merged = f"{val3_str}{val4_str}"
            elif val3_str:
                merged = val3_str
            elif val4_str:
                merged = val4_str
            else:
                merged = ""
            
            print(f"  列{col_idx} ({col_letter}): 主='{val3_str}' | 子='{val4_str}' | 合并='{merged}'")
    
    # 特别关注用户提到的列
    print("\n用户特别关注的列:")
    print("-" * 100)
    for col_idx in [8, 11, 14]:
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        val3_merged = get_merged_value(ws, 3, col_idx)
        val3 = ws.cell(row=3, column=col_idx).value
        val3_final = val3_merged if val3_merged is not None else val3
        
        val4_merged = get_merged_value(ws, 4, col_idx)
        val4 = ws.cell(row=4, column=col_idx).value
        val4_final = val4_merged if val4_merged is not None else val4
        
        val3_str = str(val3_final).strip() if val3_final else ""
        val4_str = str(val4_final).strip() if val4_final else ""
        merged = f"{val3_str}{val4_str}" if (val3_str and val4_str) else (val3_str or val4_str)
        
        print(f"\n列{col_idx} ({col_letter}):")
        print(f"  第3行值: '{val3_str}'")
        print(f"  第4行值: '{val4_str}'")
        print(f"  合并结果: '{merged}'")
    
    wb.close()

if __name__ == "__main__":
    main()
