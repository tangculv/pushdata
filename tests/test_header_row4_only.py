#!/usr/bin/env python3
"""
测试：只读取第4行作为表头（修复后的逻辑）
"""

import openpyxl
from pathlib import Path

def get_merged_value(ws, row, col):
    """获取合并单元格的值"""
    if not hasattr(ws, "merged_cells"):
        return None
    for merged_range in ws.merged_cells.ranges:
        if (
            merged_range.min_row <= row <= merged_range.max_row
            and merged_range.min_col <= col <= merged_range.max_col
        ):
            return ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
    return None

def main():
    project_root = Path(__file__).resolve().parents[1]
    excel_file = project_root / "data" / "山禾田_会员储值消费分析表_2026-01-24 12_17_45_a049681_1769228269877.xlsx"
    
    wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    
    print("=" * 100)
    print("修复后的表头解析：只读取第4行（不合并第3行）")
    print("=" * 100)
    
    # 只读取第4行作为表头（修复后的逻辑）
    row4 = list(ws.iter_rows(min_row=4, max_row=4, max_col=30))[0]
    
    headers = []
    for col_idx, cell in enumerate(row4, start=1):
        # 检查合并单元格
        merged_val = get_merged_value(ws, 4, col_idx)
        val = merged_val if merged_val is not None else cell.value
        if val is not None:
            headers.append((col_idx, str(val).strip()))
    
    print(f"\n表头总数: {len(headers)}")
    print("\n所有表头列表（只读取第4行）:")
    print("-" * 100)
    
    for idx, (col_idx, header) in enumerate(headers, start=1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        print(f"{idx:3d}. 列{col_idx} ({col_letter}): {header}")
    
    print("-" * 100)
    print(f"\n共 {len(headers)} 个表头")
    
    # 验证用户特别关注的列
    print("\n" + "=" * 100)
    print("验证用户特别关注的列（修复后应该是第4行的内容）:")
    print("-" * 100)
    
    # 列8 (H) 应该是 "储值余额累计（元）"（不是"累计储值金额（元）储值余额累计（元）"）
    # 列11 (K) 应该是 "储值余额（元）"（不是"会员余额（元）储值余额（元）"）
    # 列14 (N) 应该是 "未消费储值余额占比"（不是"未消费储值占比未消费储值余额占比"）
    
    expected = {
        8: "储值余额累计（元）",
        11: "储值余额（元）",
        14: "未消费储值余额占比"
    }
    
    for col_idx, expected_header in expected.items():
        found = False
        for cidx, header in headers:
            if cidx == col_idx:
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                if header == expected_header:
                    print(f"✓ 列{col_idx} ({col_letter}): '{header}' (正确 - 只有第4行的内容)")
                else:
                    print(f"✗ 列{col_idx} ({col_letter}): 期望 '{expected_header}', 实际 '{header}'")
                found = True
                break
        if not found:
            print(f"✗ 列{col_idx}: 未找到")
    
    wb.close()

if __name__ == "__main__":
    main()
