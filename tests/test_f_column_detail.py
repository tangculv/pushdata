#!/usr/bin/env python3
"""
详细检查F列的情况
"""

import openpyxl
from pathlib import Path

def main():
    project_root = Path(__file__).resolve().parents[1]
    excel_file = project_root / "data" / "山禾田_会员储值消费分析表_2026-01-24 12_17_45_a049681_1769228269877.xlsx"
    
    wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    
    print("=" * 100)
    print("详细检查F列（总会员数）的情况")
    print("=" * 100)
    
    # 检查F列的所有行（1-10行）
    print("\nF列的所有行内容:")
    print("-" * 100)
    for row_idx in range(1, 11):
        cell = ws.cell(row=row_idx, column=6)  # F列是第6列
        val = cell.value
        val_str = str(val) if val is not None else "(空)"
        print(f"  第{row_idx}行: {val_str}")
    
    # 检查所有涉及F列的合并单元格
    print("\n涉及F列的合并单元格:")
    print("-" * 100)
    if hasattr(ws, "merged_cells"):
        found = False
        for merged_range in ws.merged_cells.ranges:
            if merged_range.min_col <= 6 <= merged_range.max_col:
                min_col_letter = openpyxl.utils.get_column_letter(merged_range.min_col)
                max_col_letter = openpyxl.utils.get_column_letter(merged_range.max_col)
                min_row_val = ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
                print(f"  行{merged_range.min_row}-{merged_range.max_row}, 列{merged_range.min_col}-{merged_range.max_col} ({min_col_letter}-{max_col_letter}): '{min_row_val}'")
                found = True
        if not found:
            print("  没有找到涉及F列的合并单元格")
    
    # 检查A-G列的情况
    print("\n" + "=" * 100)
    print("检查A-G列（第3行和第4行）:")
    print("-" * 100)
    for col_idx in range(1, 8):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        val3 = ws.cell(row=3, column=col_idx).value
        val4 = ws.cell(row=4, column=col_idx).value
        val3_str = str(val3).strip() if val3 is not None else "(空)"
        val4_str = str(val4).strip() if val4 is not None else "(空)"
        print(f"  列{col_letter} (列{col_idx}): 第3行='{val3_str}' | 第4行='{val4_str}'")
    
    # 检查第5行的数据（验证哪些列有数据）
    print("\n" + "=" * 100)
    print("第5行的数据（验证哪些列有数据）:")
    print("-" * 100)
    row5 = list(ws.iter_rows(min_row=5, max_row=5, max_col=25))[0]
    for col_idx, cell in enumerate(row5, start=1):
        val = cell.value
        if val is not None:
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            print(f"  列{col_letter} (列{col_idx}): {val}")
    
    wb.close()

if __name__ == "__main__":
    main()
