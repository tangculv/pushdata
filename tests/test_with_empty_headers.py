#!/usr/bin/env python3
"""
测试：保留空表头，验证所有列都能被识别
"""

import openpyxl
from pathlib import Path

def split_merged_cells(ws, row_idx: int, col_idx: int):
    """拆分合并单元格，返回合并区域左上角的值"""
    if not hasattr(ws, "merged_cells"):
        return None
    
    merged_ranges = list(ws.merged_cells.ranges)
    for merged_range in merged_ranges:
        if (
            merged_range.min_row <= row_idx <= merged_range.max_row
            and merged_range.min_col <= col_idx <= merged_range.max_col
        ):
            top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
            return top_left_cell.value
    return None

def fill_header_row_with_empty(ws, header_row_1based: int, max_col=200):
    """填充表头行，保留空表头"""
    header_row = list(ws.iter_rows(min_row=header_row_1based, max_row=header_row_1based, max_col=max_col))[0]
    header_values = []
    
    for col_idx, cell in enumerate(header_row, start=1):
        merged_val = split_merged_cells(ws, header_row_1based, col_idx)
        if merged_val is not None:
            header_values.append(str(merged_val).strip())
        else:
            val = cell.value
            if val is not None:
                header_values.append(str(val).strip())
            else:
                header_values.append("")  # 保留空字符串
    
    # 移除尾部连续的空列
    while header_values and not header_values[-1]:
        header_values.pop()
    
    return header_values  # 返回所有表头，包括空字符串

def main():
    project_root = Path(__file__).resolve().parents[1]
    excel_file = project_root / "data" / "山禾田_会员储值消费分析表_2026-01-24 12_17_45_a049681_1769228269877.xlsx"
    
    wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    
    print("=" * 100)
    print("测试：保留空表头，验证所有列都能被识别")
    print("=" * 100)
    
    header_row_1based = 4
    headers = fill_header_row_with_empty(ws, header_row_1based)
    
    print(f"\n表头总数（包括空表头）: {len(headers)}")
    print("\n所有表头列表（包括空表头）:")
    print("-" * 100)
    
    for idx, header in enumerate(headers, start=1):
        col_letter = openpyxl.utils.get_column_letter(idx)
        if header:
            print(f"列{idx:2d} ({col_letter:2s}): '{header}'")
        else:
            print(f"列{idx:2d} ({col_letter:2s}): (空)")
    
    # 检查用户框起来的列
    print("\n" + "=" * 100)
    print("检查用户框起来的列:")
    print("-" * 100)
    
    target_columns = {
        'F': 6,   # 总会员数
        'H': 8,   # 储值余额累计（元）
        'I': 9,   # 赠送余额累计（元）
        'N': 14,  # 未消费储值余额占比
        'O': 15,  # 未消费赠送余额占比
        'P': 16   # 合计占比
    }
    
    for col_letter, col_idx in target_columns.items():
        if col_idx <= len(headers):
            header = headers[col_idx - 1]  # col_idx是1-based，列表是0-based
            if header:
                print(f"✓ 列{col_letter} (列{col_idx}): '{header}'")
            else:
                print(f"✗ 列{col_letter} (列{col_idx}): (空) - 但第5行有数据")
        else:
            print(f"✗ 列{col_letter} (列{col_idx}): 超出表头范围")
    
    # 检查第5行的数据
    print("\n" + "=" * 100)
    print("第5行的数据（验证哪些列有数据）:")
    print("-" * 100)
    row5 = list(ws.iter_rows(min_row=5, max_row=5, max_col=len(headers)))[0]
    for col_idx, cell in enumerate(row5, start=1):
        val = cell.value
        if val is not None:
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            header = headers[col_idx - 1] if col_idx <= len(headers) else ""
            header_str = f"'{header}'" if header else "(空表头)"
            print(f"  列{col_letter} (列{col_idx}): 数据={val}, 表头={header_str}")
    
    wb.close()

if __name__ == "__main__":
    main()
