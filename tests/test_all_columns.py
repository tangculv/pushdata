#!/usr/bin/env python3
"""
检查所有列的情况，包括A-G列
"""

import openpyxl
from pathlib import Path

def split_merged_cells(ws, row_idx: int, col_idx: int):
    """拆分合并单元格，返回合并区域左上角的值"""
    if not hasattr(ws, "merged_cells"):
        return None
    
    merged_ranges = list(ws.merged_cells.ranges)
    for merged_range in merged_ranges:
        if (
            merged_range.min_row <= row_idx <= merged_range.max_row
            and merged_range.min_col <= col_idx <= merged_range.max_col
        ):
            top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
            return top_left_cell.value
    return None

def main():
    project_root = Path(__file__).resolve().parents[1]
    excel_file = project_root / "data" / "山禾田_会员储值消费分析表_2026-01-24 12_17_45_a049681_1769228269877.xlsx"
    
    wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    
    print("=" * 100)
    print("检查所有列的情况（第3行和第4行）")
    print("=" * 100)
    
    header_row_1based = 4
    
    print(f"\n表头行: 第{header_row_1based}行 (1-based)")
    print("\n逐列分析（列1-21）:")
    print("-" * 100)
    
    row3 = list(ws.iter_rows(min_row=3, max_row=3, max_col=30))[0]
    row4 = list(ws.iter_rows(min_row=4, max_row=4, max_col=30))[0]
    
    headers = []
    for col_idx in range(1, 22):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        
        # 第3行的值
        val3 = ws.cell(row=3, column=col_idx).value
        val3_str = str(val3).strip() if val3 is not None else ""
        
        # 第4行的值（检查合并单元格）
        merged_val = split_merged_cells(ws, header_row_1based, col_idx)
        val4 = ws.cell(row=4, column=col_idx).value
        val4_final = merged_val if merged_val is not None else val4
        val4_str = str(val4_final).strip() if val4_final is not None else ""
        
        # 确定表头应该是什么（按照修复后的逻辑）
        if merged_val is not None:
            # 在合并区域内，使用合并区域左上角的值
            header = str(merged_val).strip()
            note = f"（合并单元格，取左上角值）"
        elif val4 is not None:
            # 不在合并区域内，使用单元格本身的值
            header = val4_str
            note = "（第4行有值）"
        else:
            # 第4行为空，不向上查找
            header = ""
            note = "（第4行为空，不向上取）"
        
        if header or val3_str:  # 显示有内容的列
            print(f"列{col_idx:2d} ({col_letter:2s}): 第3行='{val3_str}' | 第4行='{val4_str}' | 表头='{header}' {note}")
            if header:
                headers.append((col_idx, header))
    
    print("\n" + "=" * 100)
    print(f"最终表头列表（共{len(headers)}个）:")
    print("-" * 100)
    for idx, (col_idx, header) in enumerate(headers, start=1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        print(f"{idx:3d}. 列{col_idx} ({col_letter}): {header}")
    
    print("\n" + "=" * 100)
    print("说明:")
    print("- A-G列（列1-7）在第4行是空的，按照修复后的逻辑，这些列的表头应该是空的")
    print("- 这是符合要求的：'以数据区上方一行作为表头，去除一切向上取的操作和逻辑'")
    print("- 如果业务需要这些列的表头，需要在Excel文件中将这些表头放在第4行")
    
    wb.close()

if __name__ == "__main__":
    main()
