#!/usr/bin/env python3
"""
检查所有合并单元格，特别是跨行的合并单元格
"""

import openpyxl
from pathlib import Path

def main():
    project_root = Path(__file__).resolve().parents[1]
    excel_file = project_root / "data" / "山禾田_会员储值消费分析表_2026-01-24 12_17_45_a049681_1769228269877.xlsx"
    
    wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    
    print("=" * 100)
    print("检查所有合并单元格（特别是跨行的）")
    print("=" * 100)
    
    if not hasattr(ws, "merged_cells"):
        print("工作表没有合并单元格")
        wb.close()
        return
    
    merged_ranges = list(ws.merged_cells.ranges)
    
    if not merged_ranges:
        print("没有找到合并单元格")
        wb.close()
        return
    
    print(f"\n找到 {len(merged_ranges)} 个合并单元格区域")
    print("\n所有合并单元格区域:")
    print("-" * 100)
    
    # 分类显示
    cross_row_merges = []  # 跨行的合并（涉及第3行和第4行）
    same_row_merges = []   # 同行的合并
    
    for merged_range in merged_ranges:
        min_col_letter = openpyxl.utils.get_column_letter(merged_range.min_col)
        max_col_letter = openpyxl.utils.get_column_letter(merged_range.max_col)
        min_row_val = ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
        val_str = str(min_row_val).strip() if min_row_val is not None else "(空)"
        
        if merged_range.min_row != merged_range.max_row:
            # 跨行合并
            if merged_range.min_row <= 4 and merged_range.max_row >= 3:
                cross_row_merges.append((merged_range, val_str, min_col_letter, max_col_letter))
            print(f"  跨行: 行{merged_range.min_row}-{merged_range.max_row}, 列{merged_range.min_col}-{merged_range.max_col} ({min_col_letter}-{max_col_letter}): '{val_str}'")
        else:
            # 同行合并
            same_row_merges.append((merged_range, val_str, min_col_letter, max_col_letter))
            if merged_range.min_row <= 4:
                print(f"  同行: 行{merged_range.min_row}, 列{merged_range.min_col}-{merged_range.max_col} ({min_col_letter}-{max_col_letter}): '{val_str}'")
    
    print("\n" + "=" * 100)
    print("特别关注：涉及第3行和第4行的跨行合并单元格")
    print("-" * 100)
    if cross_row_merges:
        for merged_range, val_str, min_col, max_col in cross_row_merges:
            print(f"  行{merged_range.min_row}-{merged_range.max_row}, 列{merged_range.min_col}-{merged_range.max_col} ({min_col}-{max_col}): '{val_str}'")
            
            # 检查这个合并区域覆盖了哪些列
            for col_idx in range(merged_range.min_col, merged_range.max_col + 1):
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                val4 = ws.cell(row=4, column=col_idx).value
                print(f"    - 列{col_letter} (列{col_idx}): 第4行值='{val4}' (应该从合并区域左上角获取: '{val_str}')")
    else:
        print("  没有找到涉及第3行和第4行的跨行合并单元格")
        print("  这意味着A-G列在第4行是空的，无法通过合并单元格获取表头")
    
    # 检查A-G列是否有跨行合并
    print("\n" + "=" * 100)
    print("检查A-G列是否有跨行合并（F3-F4等）:")
    print("-" * 100)
    for col_idx in range(1, 8):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        found_merge = False
        for merged_range in merged_ranges:
            if (
                merged_range.min_col <= col_idx <= merged_range.max_col
                and merged_range.min_row <= 4 <= merged_range.max_row
                and merged_range.min_row < merged_range.max_row  # 跨行
            ):
                min_col_letter = openpyxl.utils.get_column_letter(merged_range.min_col)
                max_col_letter = openpyxl.utils.get_column_letter(merged_range.max_col)
                min_row_val = ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
                print(f"  列{col_letter} (列{col_idx}): 在合并区域内 (行{merged_range.min_row}-{merged_range.max_row}, 列{min_col_letter}-{max_col_letter}): '{min_row_val}'")
                found_merge = True
                break
        if not found_merge:
            val3 = ws.cell(row=3, column=col_idx).value
            val4 = ws.cell(row=4, column=col_idx).value
            print(f"  列{col_letter} (列{col_idx}): 不在合并区域内，第3行='{val3}', 第4行='{val4}'")
    
    wb.close()

if __name__ == "__main__":
    main()
