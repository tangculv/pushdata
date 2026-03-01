#!/usr/bin/env python3
"""
分析表头解析问题的根本原因
"""

import openpyxl
from pathlib import Path

def main():
    project_root = Path(__file__).resolve().parents[1]
    excel_file = project_root / "data" / "山禾田_会员储值消费分析表_2026-01-24 12_17_45_a049681_1769228269877.xlsx"
    
    wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    
    print("=" * 100)
    print("问题根源分析：会员储值消费分析表的表头解析")
    print("=" * 100)
    
    print("\n【问题描述】")
    print("1. 表头行已经确定是第4行（1-based）")
    print("2. 但代码把第3行的'累计储值金额（元）'合并到了第4行的表头中")
    print("3. 导致H4列变成了'累计储值金额（元）储值余额累计（元）'")
    
    print("\n【当前代码逻辑】")
    print("- 在 excel_detect.py 第224-228行：")
    print("  if spec.file_type == FILETYPE_MEMBER_STORAGE:")
    print("      main_header_row_1based = header_row_1based - 1  # 第3行（主表头）")
    print("      next_row_1based = header_row_1based  # 第4行（子表头）")
    print("- 然后调用 fill_header_row(ws, 3, 4) 把第3行和第4行合并")
    
    print("\n【实际情况】")
    print("- 第3行（H3）: '累计储值金额（元）' - 这是一个合并单元格，横跨H、I、J三列")
    print("- 第4行（H4）: '储值余额累计（元）' - 这是真正的表头")
    print("- 第4行（I4）: '赠送余额累计（元）' - 这是真正的表头")
    print("- 第4行（J4）: '合计（元）' - 这是真正的表头")
    
    print("\n【实际表头内容（第4行）】")
    print("-" * 100)
    row4 = list(ws.iter_rows(min_row=4, max_row=4, max_col=30))[0]
    for col_idx, cell in enumerate(row4, start=1):
        val = cell.value
        if val is not None:
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            print(f"  列{col_idx} ({col_letter}): '{val}'")
    
    print("\n【第3行的内容（分类标题，不应合并到表头）】")
    print("-" * 100)
    row3 = list(ws.iter_rows(min_row=3, max_row=3, max_col=30))[0]
    for col_idx, cell in enumerate(row3, start=1):
        val = cell.value
        if val is not None:
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            # 检查是否是合并单元格
            merged_ranges = list(ws.merged_cells.ranges) if hasattr(ws, "merged_cells") else []
            is_merged = False
            merge_span = ""
            for merged_range in merged_ranges:
                if (
                    merged_range.min_row <= 3 <= merged_range.max_row
                    and merged_range.min_col <= col_idx <= merged_range.max_col
                ):
                    is_merged = True
                    if merged_range.min_col != merged_range.max_col:
                        merge_span = f" (合并列 {openpyxl.utils.get_column_letter(merged_range.min_col)}-{openpyxl.utils.get_column_letter(merged_range.max_col)})"
                    break
            print(f"  列{col_idx} ({col_letter}): '{val}'{merge_span}")
    
    print("\n【根本原因】")
    print("-" * 100)
    print("代码错误地认为第3行是'主表头'，第4行是'子表头'，需要合并")
    print("但实际上：")
    print("  - 第4行才是真正的表头行（已经确定）")
    print("  - 第3行只是视觉上的分类标题，用于组织多个相关列")
    print("  - 第3行的'累计储值金额（元）'横跨H、I、J三列，表示这三列都属于'累计储值金额'这个分类")
    print("  - 但表头本身应该是第4行的内容，不应该合并第3行")
    
    print("\n【正确的处理方式】")
    print("-" * 100)
    print("对于会员储值消费分析表：")
    print("  - 表头行就是第4行，不需要向上合并第3行")
    print("  - 第3行的内容只是视觉分类，不影响表头解析")
    print("  - 应该直接读取第4行作为表头，不调用 fill_header_row 的向上合并逻辑")
    
    print("\n【需要修改的代码位置】")
    print("-" * 100)
    print("siyu_etl/excel_detect.py 第224-228行")
    print("应该改为：")
    print("  if spec.file_type == FILETYPE_MEMBER_STORAGE:")
    print("      # 表头行就是第4行，不需要合并第3行")
    print("      main_header_row_1based = header_row_1based  # 第4行")
    print("      next_row_1based = None  # 不需要向下合并")
    
    wb.close()

if __name__ == "__main__":
    main()
