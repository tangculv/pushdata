#!/usr/bin/env python3
"""
测试脚本：读取会员储值消费分析表并列出所有表头
直接使用openpyxl，不依赖ETL模块
"""

import openpyxl
from pathlib import Path

def split_merged_cells(ws, row_idx: int, col_idx: int):
    """拆分合并单元格，返回合并区域左上角的值"""
    if not hasattr(ws, "merged_cells"):
        return None
    
    merged_ranges = list(ws.merged_cells.ranges)
    for merged_range in merged_ranges:
        if (
            merged_range.min_row <= row_idx <= merged_range.max_row
            and merged_range.min_col <= col_idx <= merged_range.max_col
        ):
            top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
            return top_left_cell.value
    return None

def fill_header_row(ws, header_row_1based: int, next_row_1based=None, max_col=200):
    """填充表头行，处理合并单元格和多行表头"""
    header_row = list(ws.iter_rows(min_row=header_row_1based, max_row=header_row_1based, max_col=max_col))[0]
    header_values = []
    
    # Step 1: 读取主表头行
    for col_idx, cell in enumerate(header_row, start=1):
        merged_val = split_merged_cells(ws, header_row_1based, col_idx)
        if merged_val is not None:
            header_values.append(str(merged_val).strip() if merged_val else "")
        else:
            val = cell.value
            header_values.append(str(val).strip() if val is not None else "")
    
    # Step 2: 向上查找填充空值
    for col_idx in range(len(header_values)):
        if not header_values[col_idx]:
            for check_row in range(header_row_1based - 1, 0, -1):
                merged_val = split_merged_cells(ws, check_row, col_idx + 1)
                if merged_val is not None:
                    header_values[col_idx] = str(merged_val).strip()
                    break
                check_cell = ws.cell(row=check_row, column=col_idx + 1)
                if check_cell.value is not None:
                    header_values[col_idx] = str(check_cell.value).strip()
                    break
    
    # Step 3: 处理多行表头（向下合并）
    if next_row_1based is not None:
        next_row = list(ws.iter_rows(min_row=next_row_1based, max_row=next_row_1based, max_col=max_col))[0]
        max_cols = max(len(header_values), len(next_row))
        while len(header_values) < max_cols:
            header_values.append("")
        
        for col_idx in range(max_cols):
            merged_val = split_merged_cells(ws, next_row_1based, col_idx + 1)
            if merged_val is not None:
                next_cell_val = merged_val
            else:
                next_cell_val = next_row[col_idx].value if col_idx < len(next_row) else None
            
            next_val = str(next_cell_val).strip() if next_cell_val is not None else ""
            if not header_values[col_idx] and next_val:
                header_values[col_idx] = next_val
            elif header_values[col_idx] and next_val:
                header_values[col_idx] = f"{header_values[col_idx]}{next_val}"
    
    # 移除尾部空列
    while header_values and not header_values[-1]:
        header_values.pop()
    
    return [h for h in header_values if h]  # 只返回非空表头

def main():
    # Excel文件路径
    project_root = Path(__file__).resolve().parents[1]
    excel_file = project_root / "data" / "山禾田_会员储值消费分析表_2026-01-24 12_17_45_a049681_1769228269877.xlsx"
    
    if not excel_file.exists():
        print(f"错误：文件不存在: {excel_file}")
        return
    
    print(f"正在读取文件: {excel_file.name}")
    print("=" * 80)
    
    try:
        wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        
        # 会员储值消费分析表：第3行是主表头，第4行是子表头
        main_header_row = 3  # 主表头行（1-based）
        sub_header_row = 4   # 子表头行（1-based）
        
        print(f"\n主表头行: 第{main_header_row}行")
        print(f"子表头行: 第{sub_header_row}行")
        print(f"数据起始行: 第{sub_header_row + 1}行")
        
        # 读取并合并表头
        headers = fill_header_row(ws, main_header_row, sub_header_row)
        
        print(f"\n表头总数: {len(headers)}")
        print("\n所有表头列表:")
        print("-" * 80)
        
        # 列出所有表头，带索引
        for idx, header in enumerate(headers, start=1):
            print(f"{idx:3d}. {header}")
        
        print("-" * 80)
        print(f"\n共 {len(headers)} 个表头")
        
        # 显示原始行内容（用于调试）
        print("\n" + "=" * 80)
        print("原始表头行内容（用于参考）:")
        print("-" * 80)
        
        print(f"\n第{main_header_row}行（主表头）:")
        main_row = list(ws.iter_rows(min_row=main_header_row, max_row=main_header_row, max_col=50))[0]
        main_values = [str(cell.value).strip() if cell.value is not None else "" for cell in main_row]
        for idx, val in enumerate(main_values, start=1):
            if val:
                print(f"  列{idx}: {val}")
        
        print(f"\n第{sub_header_row}行（子表头）:")
        sub_row = list(ws.iter_rows(min_row=sub_header_row, max_row=sub_header_row, max_col=50))[0]
        sub_values = [str(cell.value).strip() if cell.value is not None else "" for cell in sub_row]
        for idx, val in enumerate(sub_values, start=1):
            if val:
                print(f"  列{idx}: {val}")
        
        wb.close()
        
    except Exception as e:
        print(f"错误: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
