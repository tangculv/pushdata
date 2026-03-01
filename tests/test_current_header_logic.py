#!/usr/bin/env python3
"""
按照当前修复后的逻辑，打印识别到的表头
"""

import openpyxl
from pathlib import Path

def split_merged_cells(ws, row_idx: int, col_idx: int):
    """拆分合并单元格，返回合并区域左上角的值"""
    if not hasattr(ws, "merged_cells"):
        return None
    
    merged_ranges = list(ws.merged_cells.ranges)
    for merged_range in merged_ranges:
        if (
            merged_range.min_row <= row_idx <= merged_range.max_row
            and merged_range.min_col <= col_idx <= merged_range.max_col
        ):
            top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
            return top_left_cell.value
    return None

def fill_header_row_current_logic(ws, header_row_1based: int, max_col=200):
    """
    当前修复后的逻辑：
    1. 数据区是表头行的下一行开始
    2. 对于数据区上方的区域（包括表头行），如果单元格在合并区域内，取合并区域左上角的值
    3. 以指定的表头行作为表头，不进行向上查找和向下合并
    """
    header_row = list(ws.iter_rows(min_row=header_row_1based, max_row=header_row_1based, max_col=max_col))[0]
    header_values = []
    
    print(f"\n处理表头行第{header_row_1based}行的每个单元格:")
    print("-" * 100)
    
    # 读取表头行的每个单元格
    for col_idx, cell in enumerate(header_row, start=1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        
        # 检查当前单元格是否在合并区域内
        merged_val = split_merged_cells(ws, header_row_1based, col_idx)
        
        if merged_val is not None:
            # 如果在合并区域内，使用合并区域左上角的值
            header_value = str(merged_val).strip()
            header_values.append(header_value)
            print(f"  列{col_idx:2d} ({col_letter:2s}): 在合并区域内 -> 取左上角值 = '{header_value}'")
        else:
            # 如果不在合并区域内，使用单元格本身的值
            val = cell.value
            if val is not None:
                header_value = str(val).strip()
                header_values.append(header_value)
                print(f"  列{col_idx:2d} ({col_letter:2s}): 不在合并区域内 -> 单元格值 = '{header_value}'")
            else:
                # 空值保留为空字符串，不向上查找
                header_values.append("")
                print(f"  列{col_idx:2d} ({col_letter:2s}): 不在合并区域内 -> 单元格值为空 -> 保留为空字符串（不向上查找）")
    
    # 移除尾部连续的空列
    print("\n移除尾部连续的空列:")
    print("-" * 100)
    original_count = len(header_values)
    while header_values and not header_values[-1]:
        removed = header_values.pop()
        print(f"  移除尾部空列: '{removed}'")
    print(f"  移除前: {original_count} 列, 移除后: {len(header_values)} 列")
    
    return header_values

def main():
    project_root = Path(__file__).resolve().parents[1]
    excel_file = project_root / "data" / "山禾田_会员储值消费分析表_2026-01-24 12_17_45_a049681_1769228269877.xlsx"
    
    if not excel_file.exists():
        print(f"错误：文件不存在: {excel_file}")
        return
    
    print("=" * 100)
    print("按照当前修复后的逻辑识别表头")
    print("=" * 100)
    print(f"\n文件: {excel_file.name}")
    
    # 注意：read_only模式下无法读取合并单元格信息，需要使用read_only=False
    wb = openpyxl.load_workbook(excel_file, read_only=False, data_only=True, keep_vba=False)
    ws = wb.worksheets[0]
    
    # 会员储值消费分析表：表头行是第4行（1-based）
    header_row_1based = 4
    data_start_row = header_row_1based + 1  # 第5行开始是数据区
    
    print(f"\n表头行: 第{header_row_1based}行 (1-based)")
    print(f"数据区起始行: 第{data_start_row}行 (1-based)")
    
    # 使用当前逻辑读取表头
    headers = fill_header_row_current_logic(ws, header_row_1based)
    
    print("\n" + "=" * 100)
    print("最终识别到的表头列表:")
    print("-" * 100)
    print(f"\n表头总数: {len(headers)}")
    print("\n所有表头（逐列显示）:")
    print("-" * 100)
    
    for idx, header in enumerate(headers, start=1):
        col_letter = openpyxl.utils.get_column_letter(idx)
        if header:
            print(f"{idx:3d}. 列{idx:2d} ({col_letter:2s}): '{header}'")
        else:
            print(f"{idx:3d}. 列{idx:2d} ({col_letter:2s}): (空)")
    
    print("-" * 100)
    
    # 统计
    non_empty_count = sum(1 for h in headers if h)
    empty_count = len(headers) - non_empty_count
    print(f"\n统计:")
    print(f"  非空表头: {non_empty_count} 个")
    print(f"  空表头: {empty_count} 个")
    print(f"  总计: {len(headers)} 个")
    
    # 检查用户框起来的列
    print("\n" + "=" * 100)
    print("检查用户框起来的列（F、H、I、N、O、P）:")
    print("-" * 100)
    
    target_columns = {
        'F': 6,   # 总会员数
        'H': 8,   # 储值余额累计（元）
        'I': 9,   # 赠送余额累计（元）
        'N': 14,  # 未消费储值余额占比
        'O': 15,  # 未消费赠送余额占比
        'P': 16   # 合计占比
    }
    
    for col_letter, col_idx in target_columns.items():
        if col_idx <= len(headers):
            header = headers[col_idx - 1]  # col_idx是1-based，列表是0-based
            if header:
                print(f"✓ 列{col_letter} (列{col_idx}): '{header}'")
            else:
                print(f"✗ 列{col_letter} (列{col_idx}): (空) - 但第5行有数据")
        else:
            print(f"✗ 列{col_letter} (列{col_idx}): 超出表头范围（表头只有{len(headers)}列）")
    
    wb.close()

if __name__ == "__main__":
    main()
