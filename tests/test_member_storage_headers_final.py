#!/usr/bin/env python3
"""
最终版本：列出会员储值消费分析表的所有表头
"""

import openpyxl
from pathlib import Path

def get_merged_value(ws, row, col):
    """获取合并单元格的值"""
    if not hasattr(ws, "merged_cells"):
        return None
    for merged_range in ws.merged_cells.ranges:
        if (
            merged_range.min_row <= row <= merged_range.max_row
            and merged_range.min_col <= col <= merged_range.max_col
        ):
            return ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
    return None

def main():
    project_root = Path(__file__).resolve().parents[1]
    excel_file = project_root / "data" / "山禾田_会员储值消费分析表_2026-01-24 12_17_45_a049681_1769228269877.xlsx"
    
    if not excel_file.exists():
        print(f"错误：文件不存在: {excel_file}")
        return
    
    wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    
    print("=" * 100)
    print("会员储值消费分析表 - 所有表头")
    print("=" * 100)
    
    # 读取第3行和第4行
    row3 = list(ws.iter_rows(min_row=3, max_row=3, max_col=30))[0]
    row4 = list(ws.iter_rows(min_row=4, max_row=4, max_col=30))[0]
    
    headers_list = []
    max_col = max(len(row3), len(row4))
    
    for col_idx in range(1, max_col + 1):
        # 获取第3行的值（检查合并单元格）
        val3_merged = get_merged_value(ws, 3, col_idx)
        val3 = ws.cell(row=3, column=col_idx).value
        val3_final = val3_merged if val3_merged is not None else val3
        
        # 获取第4行的值（检查合并单元格）
        val4_merged = get_merged_value(ws, 4, col_idx)
        val4 = ws.cell(row=4, column=col_idx).value
        val4_final = val4_merged if val4_merged is not None else val4
        
        # 合并表头
        val3_str = str(val3_final).strip() if val3_final else ""
        val4_str = str(val4_final).strip() if val4_final else ""
        
        if val3_str and val4_str:
            merged = f"{val3_str}{val4_str}"
        elif val3_str:
            merged = val3_str
        elif val4_str:
            merged = val4_str
        else:
            continue  # 跳过空列
        
        headers_list.append((col_idx, val3_str, val4_str, merged))
    
    print(f"\n表头总数: {len(headers_list)}")
    print("\n所有表头列表:")
    print("-" * 100)
    
    for idx, (col_idx, val3, val4, merged) in enumerate(headers_list, start=1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        if val3 and val4:
            print(f"{idx:3d}. 列{col_idx} ({col_letter}): {merged}")
            print(f"     主表头: {val3}")
            print(f"     子表头: {val4}")
        else:
            print(f"{idx:3d}. 列{col_idx} ({col_letter}): {merged}")
    
    print("-" * 100)
    print(f"\n共 {len(headers_list)} 个表头")
    
    # 验证用户特别关注的列
    print("\n" + "=" * 100)
    print("验证用户特别关注的列:")
    print("-" * 100)
    
    target_headers = {
        8: "累计储值金额（元）储值余额累计（元）",
        11: "会员余额（元）储值余额（元）",
        14: "未消费储值占比未消费储值余额占比"
    }
    
    for col_idx, expected in target_headers.items():
        found = False
        for cidx, _, _, merged in headers_list:
            if cidx == col_idx:
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                if merged == expected:
                    print(f"✓ 列{col_idx} ({col_letter}): {merged} (正确)")
                    found = True
                else:
                    print(f"✗ 列{col_idx} ({col_letter}): 期望 '{expected}', 实际 '{merged}'")
                    found = True
                break
        if not found:
            print(f"✗ 列{col_idx}: 未找到")
    
    wb.close()

if __name__ == "__main__":
    main()
