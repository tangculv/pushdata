#!/usr/bin/env python3
"""
检查遗漏的列：F列（总会员数）等
"""

import openpyxl
from pathlib import Path

def split_merged_cells(ws, row_idx: int, col_idx: int):
    """拆分合并单元格，返回合并区域左上角的值"""
    if not hasattr(ws, "merged_cells"):
        return None
    
    merged_ranges = list(ws.merged_cells.ranges)
    for merged_range in merged_ranges:
        if (
            merged_range.min_row <= row_idx <= merged_range.max_row
            and merged_range.min_col <= col_idx <= merged_range.max_col
        ):
            top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
            return top_left_cell.value
    return None

def main():
    project_root = Path(__file__).resolve().parents[1]
    excel_file = project_root / "data" / "山禾田_会员储值消费分析表_2026-01-24 12_17_45_a049681_1769228269877.xlsx"
    
    wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    
    print("=" * 100)
    print("检查用户框起来的列：F列、H列、I列、N列、O列、P列")
    print("=" * 100)
    
    header_row_1based = 4
    
    # 检查所有合并单元格
    print("\n所有合并单元格区域（涉及第3行和第4行）:")
    print("-" * 100)
    if hasattr(ws, "merged_cells"):
        for merged_range in ws.merged_cells.ranges:
            if merged_range.min_row <= 4 and merged_range.max_row >= 3:
                min_col_letter = openpyxl.utils.get_column_letter(merged_range.min_col)
                max_col_letter = openpyxl.utils.get_column_letter(merged_range.max_col)
                min_row_val = ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
                print(f"  行{merged_range.min_row}-{merged_range.max_row}, 列{merged_range.min_col}-{merged_range.max_col} ({min_col_letter}-{max_col_letter}): '{min_row_val}'")
    
    # 检查用户关注的列
    print("\n" + "=" * 100)
    print("检查用户框起来的列:")
    print("-" * 100)
    
    target_columns = {
        'F': 6,   # 总会员数
        'H': 8,   # 储值余额累计（元）
        'I': 9,   # 赠送余额累计（元）
        'N': 14,  # 未消费储值余额占比
        'O': 15,  # 未消费赠送余额占比
        'P': 16   # 合计占比
    }
    
    for col_letter, col_idx in target_columns.items():
        print(f"\n列{col_letter} (列{col_idx}):")
        
        # 第3行的值
        val3 = ws.cell(row=3, column=col_idx).value
        val3_str = str(val3).strip() if val3 is not None else ""
        
        # 第4行的值（检查合并单元格）
        merged_val = split_merged_cells(ws, header_row_1based, col_idx)
        val4 = ws.cell(row=4, column=col_idx).value
        val4_final = merged_val if merged_val is not None else val4
        val4_str = str(val4_final).strip() if val4_final is not None else ""
        
        print(f"  第3行值: '{val3_str}'")
        print(f"  第4行值: '{val4_str}'")
        print(f"  是否在合并区域内: {'是' if merged_val is not None else '否'}")
        
        # 确定表头应该是什么（按照修复后的逻辑）
        if merged_val is not None:
            header = str(merged_val).strip()
            print(f"  表头（合并单元格左上角）: '{header}'")
        elif val4 is not None:
            header = val4_str
            print(f"  表头（第4行值）: '{header}'")
        else:
            header = ""
            print(f"  表头: '{header}' (空)")
        
        # 检查第5行的数据（验证是否有数据）
        val5 = ws.cell(row=5, column=col_idx).value
        val5_str = str(val5) if val5 is not None else ""
        print(f"  第5行数据: '{val5_str}'")
    
    # 完整读取第4行所有列
    print("\n" + "=" * 100)
    print("第4行所有列（完整读取）:")
    print("-" * 100)
    row4 = list(ws.iter_rows(min_row=4, max_row=4, max_col=30))[0]
    headers = []
    for col_idx, cell in enumerate(row4, start=1):
        # 检查合并单元格
        merged_val = split_merged_cells(ws, header_row_1based, col_idx)
        val = merged_val if merged_val is not None else cell.value
        
        if val is not None:
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            val_str = str(val).strip()
            headers.append((col_idx, col_letter, val_str))
            print(f"  列{col_idx} ({col_letter}): '{val_str}'")
    
    print(f"\n第4行共有 {len(headers)} 个非空表头")
    
    wb.close()

if __name__ == "__main__":
    main()
