#!/usr/bin/env python3
"""
验证修复后的表头解析逻辑
直接使用openpyxl模拟修复后的fill_header_row逻辑
"""

import openpyxl
from pathlib import Path

def split_merged_cells(ws, row_idx: int, col_idx: int):
    """拆分合并单元格，返回合并区域左上角的值"""
    if not hasattr(ws, "merged_cells"):
        return None
    
    merged_ranges = list(ws.merged_cells.ranges)
    for merged_range in merged_ranges:
        if (
            merged_range.min_row <= row_idx <= merged_range.max_row
            and merged_range.min_col <= col_idx <= merged_range.max_col
        ):
            top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
            return top_left_cell.value
    return None

def fill_header_row_fixed(ws, header_row_1based: int, max_col=200):
    """
    修复后的fill_header_row逻辑：
    1. 数据区是表头行的下一行开始
    2. 对于数据区上方的区域（包括表头行），如果单元格在合并区域内，取合并区域左上角的值
    3. 以指定的表头行作为表头，不进行向上查找和向下合并
    """
    header_row = list(ws.iter_rows(min_row=header_row_1based, max_row=header_row_1based, max_col=max_col))[0]
    header_values = []
    
    # 读取表头行的每个单元格
    for col_idx, cell in enumerate(header_row, start=1):
        # 检查当前单元格是否在合并区域内
        merged_val = split_merged_cells(ws, header_row_1based, col_idx)
        if merged_val is not None:
            # 如果在合并区域内，使用合并区域左上角的值
            header_values.append(str(merged_val).strip())
        else:
            # 如果不在合并区域内，使用单元格本身的值
            val = cell.value
            if val is not None:
                header_values.append(str(val).strip())
            else:
                # 空值保留为空字符串，不向上查找
                header_values.append("")
    
    # 移除尾部空列
    while header_values and not header_values[-1]:
        header_values.pop()
    
    # 只返回非空表头
    return [h for h in header_values if h]

def main():
    project_root = Path(__file__).resolve().parents[1]
    excel_file = project_root / "data" / "山禾田_会员储值消费分析表_2026-01-24 12_17_45_a049681_1769228269877.xlsx"
    
    if not excel_file.exists():
        print(f"错误：文件不存在: {excel_file}")
        return
    
    print("=" * 100)
    print("验证修复后的表头解析逻辑")
    print("=" * 100)
    
    wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    
    # 会员储值消费分析表：表头行是第4行（1-based）
    header_row_1based = 4
    data_start_row = header_row_1based + 1  # 第5行开始是数据区
    
    print(f"\n表头行: 第{header_row_1based}行 (1-based)")
    print(f"数据区起始行: 第{data_start_row}行 (1-based)")
    
    # 使用修复后的逻辑读取表头
    headers = fill_header_row_fixed(ws, header_row_1based)
    
    print(f"\n表头总数: {len(headers)}")
    print("\n所有表头列表:")
    print("-" * 100)
    
    for idx, header in enumerate(headers, start=1):
        print(f"{idx:3d}. {header}")
    
    print("-" * 100)
    print(f"\n共 {len(headers)} 个表头")
    
    # 验证关键点
    print("\n" + "=" * 100)
    print("验证关键点:")
    print("-" * 100)
    
    # 1. 验证不应该包含第3行的内容（单独出现）
    row3_keywords = ["累计储值金额（元）", "会员余额（元）", "未消费储值占比", "会员消费储值金额（元）"]
    found_issues = []
    
    for header in headers:
        for keyword in row3_keywords:
            # 如果表头只包含第3行的关键词，没有第4行的内容，说明有问题
            if keyword == header:  # 完全匹配，说明只有第3行的内容
                found_issues.append(f"表头 '{header}' 只包含第3行的内容")
            elif keyword in header and not any(
                kw in header for kw in ["储值余额累计", "赠送余额累计", "合计", "储值余额（元）", 
                                       "赠送余额（元）", "未消费储值余额占比", "消费储值余额（元）"]
            ):
                found_issues.append(f"表头 '{header}' 可能错误地合并了第3行的内容")
    
    if found_issues:
        print("✗ 发现问题：")
        for issue in found_issues:
            print(f"  - {issue}")
    else:
        print("✓ 未发现第3行内容被错误合并到表头中")
    
    # 2. 验证关键列的表头应该是第4行的内容
    print("\n关键列的表头验证:")
    expected_headers = {
        "储值余额累计（元）": "列8 (H) - 应该是第4行的内容，不包含第3行的'累计储值金额（元）'",
        "储值余额（元）": "列11 (K) - 应该是第4行的内容，不包含第3行的'会员余额（元）'",
        "未消费储值余额占比": "列14 (N) - 应该是第4行的内容，不包含第3行的'未消费储值占比'"
    }
    
    all_correct = True
    for expected_header, description in expected_headers.items():
        if expected_header in headers:
            print(f"✓ 找到 '{expected_header}' ({description})")
        else:
            print(f"✗ 未找到期望的表头 '{expected_header}' ({description})")
            # 查找类似的
            similar = [h for h in headers if any(word in h for word in expected_header.split())]
            if similar:
                print(f"  实际找到的表头: {similar}")
            all_correct = False
    
    # 3. 显示第3行和第4行的原始内容对比
    print("\n" + "=" * 100)
    print("第3行和第4行的原始内容对比（用于参考）:")
    print("-" * 100)
    
    print("\n第3行（分类标题，不应合并到表头）:")
    row3 = list(ws.iter_rows(min_row=3, max_row=3, max_col=30))[0]
    for col_idx, cell in enumerate(row3, start=1):
        val = cell.value
        if val is not None:
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            print(f"  列{col_idx} ({col_letter}): '{val}'")
    
    print("\n第4行（真正的表头行）:")
    row4 = list(ws.iter_rows(min_row=4, max_row=4, max_col=30))[0]
    for col_idx, cell in enumerate(row4, start=1):
        val = cell.value
        if val is not None:
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            print(f"  列{col_idx} ({col_letter}): '{val}'")
    
    print("\n" + "=" * 100)
    if all_correct and not found_issues:
        print("✓ 所有验证通过！修复成功！")
    else:
        print("✗ 仍有问题需要修复")
    
    wb.close()

if __name__ == "__main__":
    main()
