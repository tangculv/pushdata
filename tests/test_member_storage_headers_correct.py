#!/usr/bin/env python3
"""
正确解析会员储值消费分析表的表头 - 按列位置编号
"""

import openpyxl
from pathlib import Path

def get_merged_value(ws, row, col):
    """获取合并单元格的值"""
    if not hasattr(ws, "merged_cells"):
        return None
    for merged_range in ws.merged_cells.ranges:
        if (
            merged_range.min_row <= row <= merged_range.max_row
            and merged_range.min_col <= col <= merged_range.max_col
        ):
            return ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
    return None

def main():
    project_root = Path(__file__).resolve().parents[1]
    excel_file = project_root / "data" / "山禾田_会员储值消费分析表_2026-01-24 12_17_45_a049681_1769228269877.xlsx"
    
    if not excel_file.exists():
        print(f"错误：文件不存在: {excel_file}")
        return
    
    wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    
    print("=" * 100)
    print("会员储值消费分析表 - 所有表头（按列位置）")
    print("=" * 100)
    
    # 读取第3行和第4行
    row3 = list(ws.iter_rows(min_row=3, max_row=3, max_col=30))[0]
    row4 = list(ws.iter_rows(min_row=4, max_row=4, max_col=30))[0]
    
    headers = []
    max_col = max(len(row3), len(row4))
    
    for col_idx in range(1, max_col + 1):
        # 获取第3行的值（检查合并单元格）
        val3_merged = get_merged_value(ws, 3, col_idx)
        val3 = ws.cell(row=3, column=col_idx).value
        val3_final = val3_merged if val3_merged is not None else val3
        
        # 获取第4行的值（检查合并单元格）
        val4_merged = get_merged_value(ws, 4, col_idx)
        val4 = ws.cell(row=4, column=col_idx).value
        val4_final = val4_merged if val4_merged is not None else val4
        
        # 合并表头
        val3_str = str(val3_final).strip() if val3_final else ""
        val4_str = str(val4_final).strip() if val4_final else ""
        
        if val3_str and val4_str:
            merged = f"{val3_str}{val4_str}"
        elif val3_str:
            merged = val3_str
        elif val4_str:
            merged = val4_str
        else:
            merged = ""
        
        headers.append((col_idx, merged))
    
    # 只显示非空表头
    print("\n所有表头列表（只显示有内容的列）:")
    print("-" * 100)
    for col_idx, header in headers:
        if header:
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            print(f"列{col_idx} ({col_letter}): {header}")
    
    print("\n" + "=" * 100)
    print("按顺序编号的表头列表:")
    print("-" * 100)
    
    # 按顺序编号（只对非空表头编号）
    header_count = 0
    for col_idx, header in headers:
        if header:
            header_count += 1
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            print(f"{header_count:3d}. 列{col_idx} ({col_letter}): {header}")
    
    print("-" * 100)
    print(f"\n共 {header_count} 个表头")
    
    # 特别显示用户关注的列
    print("\n" + "=" * 100)
    print("用户特别关注的列:")
    print("-" * 100)
    for col_idx in [8, 11, 14]:
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        _, header = headers[col_idx - 1]  # col_idx是1-based，列表是0-based
        print(f"列{col_idx} ({col_letter}): {header}")
    
    wb.close()

if __name__ == "__main__":
    main()
